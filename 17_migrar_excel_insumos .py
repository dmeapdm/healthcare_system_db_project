"""
🏥 SCRIPT DE MIGRACIÓN MASIVA: Excel de Electromedicina -> tabla `inputs`
👨‍💻 Hospital Heller - Electromedicina

QUÉ HACE:
Lee las 4 hojas de la planilla histórica de insumos/repuestos/herramientas
(StockInsElect, StockRepuestos, StockInsMed, STOCK HERRAM), las limpia y
normaliza, y las carga en la tabla unificada `inputs` de tu base de datos.

ANTES DE CORRERLO:
1. Asegurate de haber corrido el script SQL "05_inputs_stock_movements_upgrade.sql"
   (crea la tabla `inputs` con la estructura correcta, incluyendo id_supplier).
2. Confirmá que tu archivo .env tiene las credenciales correctas.
3. Poné el archivo Excel en la misma carpeta que este script, o ajustá EXCEL_PATH abajo.
4. Este script es seguro de correr en modo DRY RUN (no toca la base) hasta que
   confirmes explícitamente. Leé el resumen antes de decir que sí.

CÓMO CORRERLO:
    python 17_migrar_excel_insumos.py
"""

import os
import re
from collections import Counter

import mysql.connector
import openpyxl
from dotenv import load_dotenv

load_dotenv()

# =========================================================================
# ⚙️ CONFIGURACIÓN
# =========================================================================
EXCEL_PATH = "INSUMOS___HERRAMIENTAS_DIHH_V1_0.xlsx"  # ajustá si está en otra carpeta


def conectar_base_datos():
    return mysql.connector.connect(
        host=os.getenv("DB_HOST"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        database=os.getenv("DB_NAME"),
    )


# =========================================================================
# 🧹 FUNCIONES DE LIMPIEZA DE DATOS
# =========================================================================
def clean(v):
    """Limpia texto: quita espacios extra, convierte vacíos a None."""
    if v is None:
        return None
    s = str(v).strip()
    s = re.sub(r"\s+", " ", s)
    return s if s else None


def clean_code(v):
    """Normaliza códigos internos: descarta None/'S/N'/vacíos, convierte floats (9617.0 -> '9617')."""
    if v is None:
        return None
    if isinstance(v, float):
        v = int(v)
    s = str(v).strip().upper()
    if s in ("S/N", "", "NONE"):
        return None
    return s


avisos_cantidad = []


def clean_cantidad(v, contexto=""):
    """Convierte cantidad a entero, tolerando coma decimal (formato argentino) y texto sucio."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(round(v))
    s = str(v).strip().replace(",", ".")
    try:
        val = float(s)
        redondeado = int(round(val))
        if val != redondeado:
            avisos_cantidad.append(f"{contexto}: '{v}' -> se redondeó a {redondeado}")
        return redondeado
    except ValueError:
        avisos_cantidad.append(f"{contexto}: '{v}' no es numérico -> se cargó como 0")
        return 0


codigos_generados = {"insumo_taller": 0, "repuesto_tecnico": 0, "herramienta": 0, "insumo_clinico": 0}
codigos_usados = set()
PREFIJOS = {"insumo_taller": "TALL", "repuesto_tecnico": "REP", "herramienta": "HERR", "insumo_clinico": "MED"}


def prox_codigo(tipo):
    """Genera un código interno automático (TALL-0001, REP-0001, etc.) para artículos sin código en el Excel."""
    codigos_generados[tipo] += 1
    cod = f"{PREFIJOS[tipo]}-{codigos_generados[tipo]:04d}"
    while cod in codigos_usados:
        codigos_generados[tipo] += 1
        cod = f"{PREFIJOS[tipo]}-{codigos_generados[tipo]:04d}"
    return cod


# =========================================================================
# 📥 EXTRACCIÓN Y TRANSFORMACIÓN POR HOJA
# =========================================================================
def extraer_datos(ruta_excel):
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    registros = []

    # ---- StockInsElect -> insumo_taller ----
    ws = wb["StockInsElect"]
    cat_actual = None
    for r in range(3, ws.max_row + 1):
        cantidad = ws.cell(row=r, column=3).value
        desc = clean(ws.cell(row=r, column=5).value)
        if desc is None:
            continue
        label = clean(ws.cell(row=r, column=2).value)
        if label:
            cat_actual = label
        codigo = prox_codigo("insumo_taller")
        codigos_usados.add(codigo)
        registros.append({
            "internal_code": codigo, "input_type": "insumo_taller",
            "input_category": cat_actual, "brand": "GENERICO", "model_ref": None,
            "name_input": desc,
            "unit_of_measure": (clean(ws.cell(row=r, column=4).value) or "unidad").lower(),
            "cabinet_space": clean(ws.cell(row=r, column=6).value),
            "drawer_location": clean(ws.cell(row=r, column=7).value),
            "stock": clean_cantidad(cantidad, f"StockInsElect fila {r}"),
            "unit_price": None, "id_supplier": None, "compatible_equipment": None,
            "procurement_notes": None,
        })

    # ---- StockRepuestos -> repuesto_tecnico ----
    ws = wb["StockRepuestos"]
    cat_actual = None
    for r in range(4, ws.max_row + 1):
        cantidad = ws.cell(row=r, column=3).value
        equipo_destino = clean(ws.cell(row=r, column=5).value)
        marca = clean(ws.cell(row=r, column=6).value)
        desc = clean(ws.cell(row=r, column=7).value)
        label = clean(ws.cell(row=r, column=2).value)
        if label and desc is None:
            cat_actual = label
            continue
        if desc is None:
            continue
        if label:
            cat_actual = label
        codigo = prox_codigo("repuesto_tecnico")
        codigos_usados.add(codigo)
        proveedor_txt = clean(ws.cell(row=r, column=10).value)
        gestion_txt = clean(ws.cell(row=r, column=11).value)
        notas = " | ".join(filter(None, [
            f"Proveedor (texto libre, sin vincular): {proveedor_txt}" if proveedor_txt else None,
            gestion_txt,
        ]))
        if equipo_destino:
            notas = (notas + " | " if notas else "") + f"Equipo destino original: {equipo_destino}"
        registros.append({
            "internal_code": codigo, "input_type": "repuesto_tecnico",
            "input_category": cat_actual, "brand": marca or "GENERICO", "model_ref": None,
            "name_input": desc,
            "unit_of_measure": (clean(ws.cell(row=r, column=4).value) or "unidad").lower(),
            "cabinet_space": clean(ws.cell(row=r, column=8).value),
            "drawer_location": clean(ws.cell(row=r, column=9).value),
            "stock": clean_cantidad(cantidad, f"StockRepuestos fila {r}"),
            "unit_price": None, "id_supplier": None,
            "compatible_equipment": equipo_destino,
            "procurement_notes": notas or None,
        })

    # ---- StockInsMed -> insumo_clinico ----
    ws = wb["StockInsMed"]
    for r in range(3, ws.max_row + 1):
        desc = clean(ws.cell(row=r, column=6).value)
        if desc is None:
            continue
        codigo_original = clean_code(ws.cell(row=r, column=5).value)
        if codigo_original and codigo_original not in codigos_usados:
            codigo = codigo_original
        else:
            codigo = prox_codigo("insumo_clinico")
        codigos_usados.add(codigo)
        proveedor_txt = clean(ws.cell(row=r, column=12).value)
        gestion_txt = clean(ws.cell(row=r, column=13).value)
        notas = " | ".join(filter(None, [
            f"Proveedor (texto libre, sin vincular): {proveedor_txt}" if proveedor_txt else None,
            gestion_txt,
        ]))
        cantidad = ws.cell(row=r, column=7).value
        registros.append({
            "internal_code": codigo, "input_type": "insumo_clinico",
            "input_category": clean(ws.cell(row=r, column=2).value),
            "brand": clean(ws.cell(row=r, column=3).value) or "GENERICO",
            "model_ref": clean(ws.cell(row=r, column=4).value),
            "name_input": desc,
            "unit_of_measure": (clean(ws.cell(row=r, column=8).value) or "unidad").lower(),
            "cabinet_space": clean(ws.cell(row=r, column=9).value),
            "drawer_location": clean(ws.cell(row=r, column=10).value),
            "stock": clean_cantidad(cantidad, f"StockInsMed fila {r}"),
            "unit_price": None, "id_supplier": None, "compatible_equipment": None,
            "procurement_notes": notas or None,
        })

    # ---- STOCK HERRAM -> herramienta ----
    ws = wb["STOCK HERRAM"]
    for r in range(2, ws.max_row + 1):
        desc = clean(ws.cell(row=r, column=5).value)
        if desc is None or desc.upper().startswith("TOTAL"):
            continue
        cantidad = ws.cell(row=r, column=3).value
        precio = ws.cell(row=r, column=6).value
        codigo = prox_codigo("herramienta")
        codigos_usados.add(codigo)
        registros.append({
            "internal_code": codigo, "input_type": "herramienta",
            "input_category": None, "brand": "GENERICO", "model_ref": None,
            "name_input": desc,
            "unit_of_measure": (clean(ws.cell(row=r, column=4).value) or "unidad").lower(),
            "cabinet_space": None, "drawer_location": None,
            "stock": clean_cantidad(cantidad, f"STOCK HERRAM fila {r}"),
            "unit_price": float(precio) if isinstance(precio, (int, float)) else None,
            "id_supplier": None, "compatible_equipment": None, "procurement_notes": None,
        })

    return registros


# =========================================================================
# 💾 CARGA EN LA BASE DE DATOS
# =========================================================================
def cargar_en_db(registros):
    conexion = conectar_base_datos()
    mensajero = conexion.cursor()
    insertados, errores = 0, []

    query = """
    INSERT INTO inputs (
        internal_code, input_type, input_category, brand, model_ref,
        name_input, unit_of_measure, cabinet_space, drawer_location,
        stock, unit_price, id_supplier, compatible_equipment, procurement_notes
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
    """

    try:
        for reg in registros:
            try:
                mensajero.execute(query, (
                    reg["internal_code"], reg["input_type"], reg["input_category"],
                    reg["brand"], reg["model_ref"], reg["name_input"], reg["unit_of_measure"],
                    reg["cabinet_space"], reg["drawer_location"], reg["stock"],
                    reg["unit_price"], reg["id_supplier"], reg["compatible_equipment"],
                    reg["procurement_notes"],
                ))
                insertados += 1
            except mysql.connector.Error as e:
                errores.append(f"{reg['internal_code']} ({reg['name_input'][:40]}): {e}")
        conexion.commit()
    except Exception as e:
        conexion.rollback()
        print(f"❌ Error general, se revirtió todo: {e}")
        raise
    finally:
        mensajero.close()
        conexion.close()

    return insertados, errores


# =========================================================================
# 🚀 EJECUCIÓN PRINCIPAL
# =========================================================================
if __name__ == "__main__":
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ No encontré el archivo '{EXCEL_PATH}'. Ajustá la variable EXCEL_PATH arriba del script.")
        raise SystemExit(1)

    print("📖 Leyendo y transformando la planilla...")
    registros = extraer_datos(EXCEL_PATH)

    print()
    print("=" * 60)
    print(f"📊 RESUMEN DE MIGRACIÓN — {len(registros)} artículos encontrados")
    print("=" * 60)
    for tipo, cantidad in Counter(r["input_type"] for r in registros).items():
        print(f"   {tipo}: {cantidad}")

    if avisos_cantidad:
        print()
        print(f"⚠️  {len(avisos_cantidad)} avisos de cantidad (revisá si hace falta corregir algo a mano después):")
        for a in avisos_cantidad[:20]:
            print("   -", a)
        if len(avisos_cantidad) > 20:
            print(f"   ...y {len(avisos_cantidad) - 20} más.")

    print()
    print("📋 Ejemplo de los primeros 3 registros que se van a cargar:")
    for reg in registros[:3]:
        print(f"   [{reg['internal_code']}] {reg['input_type']} | {reg['name_input'][:50]} | stock: {reg['stock']}")

    print()
    print("⚠️  Los campos 'Proveedor' del Excel se guardaron como texto en 'procurement_notes',")
    print("    NO se vincularon automáticamente a tu tabla 'supplier' (podés asignarlos a mano después).")
    print()

    respuesta = input("¿Confirmás la carga de estos registros en la base de datos? (escribí SI para continuar): ")
    if respuesta.strip().upper() == "SI":
        print("💾 Cargando en la base de datos...")
        insertados, errores = cargar_en_db(registros)
        print(f"✅ Se insertaron {insertados} de {len(registros)} registros.")
        if errores:
            print(f"⚠️  {len(errores)} registros fallaron:")
            for e in errores[:20]:
                print("   -", e)
    else:
        print("🚫 Carga cancelada. No se modificó la base de datos.")